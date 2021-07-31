'''
    Contains the objects for table extraction 
'''

from utils import (InputError, completed_qualification_valid_exams, desired_tables, detail_string,
                   escape_backslash_r, exam_results_valid_exams, is_abs_path, math_mapping, physics_mapping,
                   fm_mapping)

from pandas import isna
import os


class GradeEntry:
    '''
        Class to store the grades
    '''

    def __init__(self, qualification, subject, grade, is_predicted, year):
        self.qualification = escape_backslash_r(qualification)
        self.subject = escape_backslash_r(subject)
        self.grade = grade
        self.is_predicted = is_predicted
        self.year = year

    def __repr__(self):
        return r"Qualification: {} Subject: {} Grade: {} Year: {} Predicted {}".format(
            self.qualification, self.subject, self.grade, self.year, self.is_predicted)

    def __str__(self):
        return r"Qualification: {} Subject: {} Grade: {} Year: {} Predicted {}".format(
            self.qualification, self.subject, self.grade, self.year, self.is_predicted)


class ExtractedStudents:
    '''
        Class that stores all the students and co-ordinates the output
    '''

    def __init__(self, applicant_ids):
        self.student_ids = applicant_ids

        self.num_students = len(applicant_ids)

        self.all_students = [None]*self.num_students

        self.index = self.num_students

    def __iter__(self):
        return self

    def __next__(self):
        if self.index == 0:
            raise StopIteration

        self.index -= 1

        return self.all_students[self.index]

    def add_student_sequentially(self, new_student, counter):
        if new_student.ucas_id == self.student_ids[counter]:
            self.all_students[counter] = new_student
        else:
            raise RuntimeError("The order of adding students is incorrect \n"
                               "This should be done sequentially IDs in list should correspond")

    @staticmethod
    def populate_normal_header(ws):
        ws.cell(row=1, column=1, value="{}".format("UCAS ID"))
        ws.cell(row=1, column=2, value="{}".format("Qualification Type"))

        for i in range(0, 8, 2):
            ws.cell(row=1, column=3 + i, value="{}".format("Subject"))
            ws.cell(row=1, column=4 + i, value="{}".format("Grade"))

        return ws

    def populate_worksheet(self, desired_data, ws):
        if desired_data not in self.all_students[0].which_grades.keys():
            raise InputError(False, "Key given not found in dictionary")

        ws = self.populate_normal_header(ws)

        lb = 2

        for row_counter, student in zip(range(lb, self.num_students+lb), self.all_students):

            ws.cell(row=row_counter, column=1,
                    value="{}".format(student.ucas_id))

            if student.which_grades[desired_data]:
                ws.cell(row=row_counter, column=2, value="{}".format(
                    student.which_grades[desired_data][0].qualification))
                col_counter = 3

                for entry in student.which_grades[desired_data]:
                    ws.cell(row=row_counter, column=col_counter,
                            value="{}".format(entry.subject))
                    ws.cell(row=row_counter, column=col_counter +
                            1, value="{}".format(entry.grade))
                    col_counter += 2

        return ws

    @staticmethod
    def populate_master_header(ws):

        ws.cell(row=1, column=1, value="{}".format("UCAS ID"))
        ws.cell(row=1, column=2, value="{}".format("Qualification"))
        ws.cell(row=1, column=3, value="{}".format("Issues Importing?"))
        ws.cell(row=1, column=4, value="{}".format("Math Grade"))
        ws.cell(row=1, column=5, value="{}".format("Physics Grade"))
        ws.cell(row=1, column=10, value="{}".format("FM?"))

        for i in range(0, 4, 2):
            ws.cell(row=1, column=6 + i, value="{}".format("Subject"))
            ws.cell(row=1, column=7 + i, value="{}".format("Grade"))

        return ws

    def compile_for_master(self, ws):

        ws = self.populate_master_header(ws) 

        lb = 2

        for row_counter, student in zip(range(lb, self.num_students+lb), self.all_students):
            ws.cell(row=row_counter, column=1,
                    value="{}".format(student.ucas_id))

            is_more_than_four = False

            categorised_entries = self.sort_into_subjects(student)

            core_subjects = ["math", "physics", "fm"]

            if categorised_entries["fm"]:
                is_fm = True
                ws.cell(row=row_counter, column=10, value="Yes")
            else:
                is_fm = False
                ws.cell(row=row_counter, column=10, value="No")

            any_issues = self.log_issues(categorised_entries)
            if any_issues is None:
                ws.cell(row=row_counter, column=3, value="")
            else:
                ws.cell(row=row_counter, column=3, value=any_issues)


                            
                        

        return ws

    @staticmethod
    def log_issues(categorised_entries):
        # All lists in categorised entries are empty
        if not all(categorised_entries.values()):
            return "No valid qual & subjects."

        log = []

        for subject, entries in categorised_entries:
            if len(entries) > 1 and subject != "additional_subjects":
                log.append("Multiple {} Qual.".format(subject))
            elif len(entries) > 2:
                log.append("More than 4 subjects.")

        if len(log) > 1:
            return " ".join(log)
        elif len(log) == 1:
            return log[0]
        else:
            return None


    @staticmethod
    def sort_into_subjects(student):

        categorised_entries = {
            "math": [],
            "physics": [],
            "fm": [],
            "additional_subjects": [],
        }

        # Grade type => results/predicted/completed
        for grade_entries in student.which_grade.values():
            # List of entries are not empty
            if grade_entries:
                for entry in grade_entries:
                    if entry.subject in math_mapping().get(entry.qualification, set()):
                        categorised_entries["math"].append(entry)
                    elif entry.subject in physics_mapping().get(entry.qualification, set()):
                        categorised_entries["physics"].append(entry)
                    elif entry.subject in fm_mapping().get(entry.qualification, set()):
                        categorised_entries["fm"].append(entry)
                    else:
                        categorised_entries["additional_subjects"].append(entry)

        return categorised_entries

    def write_to_excel(self, input_abs_path):
        is_abs_path(input_abs_path)

        from openpyxl import Workbook

        wb = Workbook()

        # Create and populate workshets
        completed = wb.create_sheet("Completed Qualifications")
        completed = self.populate_worksheet("completed", completed)

        predicted = wb.create_sheet("Predicted Grades")
        predicted = self.populate_worksheet("predicted", predicted)

        exam_results = wb.create_sheet("Exam Results")
        exam_results = self.populate_worksheet("results", exam_results)

        compiled_single = wb.create_sheet("Compiled", 0)

        wb.save(os.path.join(input_abs_path, "output.xlsx"))

        return 0


class StudentGrades:
    '''
        Class for a single pdf/student
    '''

    def __init__(self, id, extracted_tables, table_headers):
        self.ucas_id = id

        self.completed_qualifications = None
        self.uncompleted_qualifications = None
        self.exam_results = None

        target_tables = desired_tables()

        for header, tbl in zip(table_headers, extracted_tables):
            if header == target_tables[0]:
                print("")
                print("Completed Qualification")
                print(tbl)
                self.completed_qualifications = tbl
            elif header == target_tables[1]:
                print("")
                print("Predicted Grades")
                print(tbl)
                self.uncompleted_qualifications = tbl
            elif header == target_tables[2]:
                print("")
                print("Exam Results")
                print(tbl)
                self.exam_results = tbl

        self.predicted_entries = []
        self.completed_entries = []
        self.results_entries = []

        self.predicted_grade_entries()
        self.examresult_entries()
        self.completed_grade_entries()

        self.which_grades = {
            "results": self.results_entries,
            "completed": self.completed_entries,
            "predicted": self.predicted_entries,
        }

    def __repr__(self):
        return "{} \n {} \n {}".format(self.completed_qualifications,
                                       self.uncompleted_qualifications,
                                       self.exam_results)

    # def handle_detailed_entry(self, input_qualification, rowCounter):
    #     target = input_qualification['Date'][rowCounter]
    #     if type(target) is str:
    #         if target in detail_string():
    #             output = []
    #             all_module_details = qualification = self.uncompleted_qualifications['Body'][row]

    #         elif type(self.uncompleted_qualifications['Date'][row]) is str:
    #             if is_pred_grade & is_grade and self.uncompleted_qualifications['Date'][row] in detail_string():
    #                 all_module_details = qualification = self.uncompleted_qualifications['Body'][row]
    #                 individual_modules = all_module_details.split("Title:")
    #                 print(individual_modules)
    #                 for module in individual_modules:
    #                     module_info = module.split("Date:")[0]
    #                     entry = GradeEntry(
    #                         None,
    #                         module_info,
    #                         None,
    #                         True,
    #                         None,
    #                     )
    #                     self.predicted_entries.append(entry)

    def completed_grade_entries(self):
        if self.completed_qualifications is None:
            return None

        for row in self.completed_qualifications.index:

            if self.is_completed_qual_valid(row):

                entry = GradeEntry(
                    self.completed_qualifications['Exam'][row],
                    self.completed_qualifications['Subject'][row],
                    self.completed_qualifications['Grade'][row],
                    False,
                    self.completed_qualifications['Date'][row].split("-")[-1],
                )

                self.completed_entries.append(entry)

        return self.completed_entries

    def is_completed_qual_valid(self, row):
        if isna(self.completed_qualifications['Exam'][row]):
            return False

        if self.completed_qualifications['Exam'][row] in completed_qualification_valid_exams():
            return True
        else:
            return False

    def examresult_entries(self):
        if self.exam_results is None:
            return None

        for row in self.exam_results.index:

            if self.is_examresult_valid(row):

                entry = GradeEntry(
                    self.exam_results['Exam Level'][row],
                    self.exam_results['Subject'][row],
                    self.exam_results['Grade'][row],
                    False,
                    self.exam_results['Date'][row].split("-")[-1],
                )

                self.results_entries.append(entry)

        return self.results_entries

    def is_examresult_valid(self, row):
        if isna(self.exam_results['Exam Level'][row]):
            return False

        if self.exam_results['Exam Level'][row] in exam_results_valid_exams():
            return True
        else:
            return False

    def predicted_grade_entries(self):
        if self.uncompleted_qualifications is None:
            return None

        for row in self.uncompleted_qualifications.index:

            is_pred_grade = isna(
                self.uncompleted_qualifications['Predicted\rGrade'][row])
            is_grade = isna(self.uncompleted_qualifications['Grade'][row])

            if is_pred_grade ^ is_grade:

                if is_pred_grade:
                    valid_grade = self.uncompleted_qualifications['Grade'][row]
                else:
                    valid_grade = self.uncompleted_qualifications['Predicted\rGrade'][row]

                if isna(self.uncompleted_qualifications['Exam'][row]):
                    qualification = self.uncompleted_qualifications['Body'][row]
                else:
                    qualification = self.uncompleted_qualifications['Exam'][row]

                entry = GradeEntry(
                    qualification,
                    self.uncompleted_qualifications['Subject'][row],
                    valid_grade,
                    True,
                    self.uncompleted_qualifications['Date'][row].split(
                        "-")[-1],
                )
                self.predicted_entries.append(entry)

            elif (not is_pred_grade) & (not is_grade):

                if "Unnamed" in self.uncompleted_qualifications['Grade'][row]:
                    valid_grade = self.uncompleted_qualifications['Predicted\rGrade'][row]
                else:
                    valid_grade = self.uncompleted_qualifications['Grade'][row]

                if isna(self.uncompleted_qualifications['Exam'][row]):
                    qualification = self.uncompleted_qualifications['Body'][row]
                else:
                    qualification = self.uncompleted_qualifications['Exam'][row]

                entry = GradeEntry(
                    qualification,
                    self.uncompleted_qualifications['Subject'][row],
                    valid_grade,
                    True,
                    self.uncompleted_qualifications['Date'][row].split(
                        "-")[-1],
                )
                self.predicted_entries.append(entry)

            elif type(self.uncompleted_qualifications['Date'][row]) is str:
                if is_pred_grade & is_grade and self.uncompleted_qualifications['Date'][row] in detail_string():
                    all_module_details = qualification = self.uncompleted_qualifications[
                        'Body'][row]
                    # Ignores the first entry which would just be the date
                    individual_modules = all_module_details.split("Title:")[1:]
                    print(individual_modules)
                    for module in individual_modules:
                        module_info = module.split("Date:")[0]
                        entry = GradeEntry(
                            None,
                            module_info,
                            None,
                            True,
                            None,
                        )
                        self.predicted_entries.append(entry)

        return self.predicted_entries
